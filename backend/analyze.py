#!/usr/bin/env python3
"""
Script d'analyse des appels Allodoct
Analyse les fichiers not_found et not_authorized pour générer des statistiques
"""

import pandas as pd
import re
from pathlib import Path
from typing import Dict, List, Tuple
import sys
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Catégories d'examens connues
CATEGORIES = {
    'IRM': ['irm', 'imagerie par résonance magnétique'],
    'SCANNER': ['scanner', 'tdm', 'tomodensitométrie', 'ct', 'coroscanner', 'angioscanner'],
    'RADIOGRAPHIE': ['radio', 'radiographie', 'rx', 'téléradiographie'],
    'MAMMOGRAPHIE': ['mammographie', 'mammo'],
    'ECHOGRAPHIE': ['échographie', 'echographie', 'écho', 'echo', 'doppler'],
    'CONE BEAM': ['cone beam', 'conebeam'],
    'DENTAIRE': ['dentaire', 'panoramique dentaire', 'orthopantomogramme']
}


def load_reference_exams(reference_file: str) -> pd.DataFrame:
    """Charge le fichier de référence des examens"""
    try:
        df = pd.read_csv(reference_file, encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(reference_file, encoding='latin-1')
    
    # Normaliser les colonnes
    df.columns = df.columns.str.strip()
    return df


def categorize_exam(exam_text: str, reference_df: pd.DataFrame) -> str:
    """
    Catégorise un examen

    Returns:
        La catégorie de l'examen
    """
    if pd.isna(exam_text) or not exam_text.strip():
        return 'INCONNU'

    exam_lower = exam_text.lower().strip()

    # Vérifier si c'est un intitulé incompris (bug du robot)
    incomprehensible_patterns = [
        r'ma mère',
        r'ma femme',
        r'mon mari',
        r'mon père',
        r'un.*pour',  # "Un scanner pour..."
        r'je veux',
        r'j\'ai besoin',
        r'\d+\s*ans',  # âge
        r'bonjour',
        r'consultation',
    ]

    for pattern in incomprehensible_patterns:
        if re.search(pattern, exam_lower):
            return 'INTITULES INCOMPRIS'

    # Déterminer la catégorie
    detected_category = 'AUTRE'
    for category, keywords in CATEGORIES.items():
        for keyword in keywords:
            if keyword in exam_lower:
                detected_category = category
                break
        if detected_category != 'AUTRE':
            break

    return detected_category


def parse_exam_identified(exam_str: str) -> List[str]:
    """Parse la colonne 'Examen Identifié' qui peut contenir plusieurs examens séparés par ';'"""
    if pd.isna(exam_str):
        return []
    return [e.strip() for e in str(exam_str).split(';') if e.strip()]


def analyze_calls(not_found_file: str, not_authorized_file: str, reference_file: str, output_file: str):
    """Analyse les fichiers d'appels et génère le rapport Excel"""
    
    print("📊 Chargement des données...")
    
    # Charger les fichiers
    df_not_found = pd.read_excel(not_found_file)
    df_not_authorized = pd.read_excel(not_authorized_file)
    reference_df = load_reference_exams(reference_file)

    # DEBUG: Afficher les colonnes des fichiers sources
    print("\n🔍 DEBUG - Colonnes du fichier not_found:")
    print(df_not_found.columns.tolist())
    print("\n🔍 DEBUG - Colonnes du fichier not_authorized:")
    print(df_not_authorized.columns.tolist())

    # CORRECTION: Supprimer la colonne 'Catégorie' si elle existe déjà dans les fichiers sources
    if 'Catégorie' in df_not_found.columns:
        print("⚠️  Suppression de la colonne 'Catégorie' existante dans not_found")
        df_not_found = df_not_found.drop(columns=['Catégorie'])
    if 'Catégorie' in df_not_authorized.columns:
        print("⚠️  Suppression de la colonne 'Catégorie' existante dans not_authorized")
        df_not_authorized = df_not_authorized.drop(columns=['Catégorie'])
    
    # Filtrer uniquement les "Transféré"
    df_not_found = df_not_found[df_not_found['Statut'] == 'Transféré'].copy()
    df_not_authorized = df_not_authorized[df_not_authorized['Statut'] == 'Transféré'].copy()
    
    print(f"   Not Found (Transféré): {len(df_not_found)} appels")
    print(f"   Not Authorized (Transféré): {len(df_not_authorized)} appels")
    
    # Ajouter une colonne de tag
    df_not_found['tag_type'] = 'exam_not_found'
    df_not_authorized['tag_type'] = 'exam_not_authorized'
    
    # Combiner les dataframes
    df_all = pd.concat([df_not_found, df_not_authorized], ignore_index=True)
    
    print("\n🔍 Analyse des examens...")

    # DEBUG: Voir quelques exemples d'examens
    print("\n🔍 DEBUG - Exemples d'examens dans 'Examen Identifié':")
    sample_exams = df_all['Examen Identifié'].head(10).tolist()
    for i, exam in enumerate(sample_exams, 1):
        print(f"  {i}. {exam}")

    # Créer une liste pour stocker les résultats détaillés
    detailed_results = []

    for idx, row in df_all.iterrows():
        exams = parse_exam_identified(row['Examen Identifié'])

        for exam in exams:
            category = categorize_exam(exam, reference_df)

            # DEBUG: Si la catégorie contient "exam_", afficher l'examen
            if 'exam_' in category.lower():
                print(f"⚠️  ANOMALIE DÉTECTÉE: Examen '{exam}' → Catégorie '{category}'")

            detailed_results.append({
                'Examen Identifié': exam,
                'Catégorie': category,
                'Tag': row['tag_type'],
                'Id Appel': row['Id'],
                'Id Externe': row['Id Externe']
            })
    
    df_detailed = pd.DataFrame(detailed_results)

    # DEBUG : Afficher toutes les catégories trouvées
    print("\n🔍 Catégories détectées:")
    print(df_detailed['Catégorie'].value_counts())

    # CORRECTION : Filtrer les catégories invalides (tags qui sont devenus catégories par erreur)
    # Les catégories valides ne doivent PAS contenir "exam_"
    invalid_categories = df_detailed[df_detailed['Catégorie'].str.contains('exam_', case=False, na=False)]
    if len(invalid_categories) > 0:
        print(f"\n⚠️  ATTENTION: {len(invalid_categories)} lignes ont des catégories invalides et seront ignorées:")
        print(invalid_categories['Catégorie'].value_counts())
        df_detailed = df_detailed[~df_detailed['Catégorie'].str.contains('exam_', case=False, na=False)]

    # Créer les statistiques
    print("\n📈 Génération des statistiques...")
    
    # Collecter tous les Id Externes par examen unique
    exam_ids_dict = {}
    for idx, row in df_detailed.iterrows():
        key = (row['Examen Identifié'], row['Catégorie'])
        if key not in exam_ids_dict:
            exam_ids_dict[key] = []
        if pd.notna(row['Id Externe']):
            exam_ids_dict[key].append(str(row['Id Externe']))
    
    stats_data = []
    
    # Ajouter 'INCONNU' aux catégories à analyser
    valid_categories = list(CATEGORIES.keys()) + ['INTITULES INCOMPRIS', 'AUTRE', 'INCONNU']

    for category in valid_categories:
        df_cat = df_detailed[df_detailed['Catégorie'] == category]

        total = len(df_cat)
        if total == 0:
            continue

        not_found = len(df_cat[df_cat['Tag'] == 'exam_not_found'])
        not_authorized = len(df_cat[df_cat['Tag'] == 'exam_not_authorized'])

        # TOUS les examens triés par fréquence avec leurs tags et ID externes
        all_exams = df_cat['Examen Identifié'].value_counts()

        # Créer une liste avec les examens, leurs tags et leurs ID externes
        exams_with_ids = []
        for exam, count in all_exams.items():
            df_exam = df_cat[df_cat['Examen Identifié'] == exam]

            # Compter les tags pour cet examen
            nf_count = len(df_exam[df_exam['Tag'] == 'exam_not_found'])
            na_count = len(df_exam[df_exam['Tag'] == 'exam_not_authorized'])

            # Récupérer tous les ID externes pour cet examen
            ids = df_exam['Id Externe'].dropna().tolist()
            # Convertir en int si possible pour éviter les .0
            ids_clean = []
            for id_val in ids:
                try:
                    # Essayer de convertir en int pour éviter les décimales
                    ids_clean.append(str(int(float(id_val))))
                except (ValueError, TypeError):
                    ids_clean.append(str(id_val))
            ids_str = '|'.join(ids_clean)  # Utiliser | comme séparateur
            exams_with_ids.append(f"{exam}§{count} (NF:{nf_count}|NA:{na_count})§{ids_str}")  # Utiliser § comme séparateur

        all_exams_str = '\n'.join(exams_with_ids)

        stats_data.append({
            'Catégorie': category,
            'Total': total,
            'exam_not_found': not_found,
            'exam_not_authorized': not_authorized,
            'Tous les examens': all_exams_str
        })
    
    df_stats = pd.DataFrame(stats_data)
    
    # Trier les examens détaillés par catégorie puis par fréquence
    exam_counts = df_detailed.groupby(['Catégorie', 'Examen Identifié']).size().reset_index(name='Occurrences')
    df_detailed_with_counts = df_detailed.merge(exam_counts, on=['Catégorie', 'Examen Identifié'])
    df_detailed_sorted = df_detailed_with_counts.sort_values(
        ['Catégorie', 'Occurrences', 'Examen Identifié'], 
        ascending=[True, False, True]
    )
    
    # Supprimer les doublons (garder une ligne par examen unique)
    df_detailed_sorted = df_detailed_with_counts.sort_values(
        ['Catégorie', 'Occurrences', 'Examen Identifié'], 
        ascending=[True, False, True]
    )
    
    # Grouper par examen unique et compter les tags séparément
    df_tag_counts = df_detailed.groupby(['Catégorie', 'Examen Identifié', 'Tag']).size().reset_index(name='count')
    df_tag_pivot = df_tag_counts.pivot_table(
        index=['Catégorie', 'Examen Identifié'],
        columns='Tag',
        values='count',
        fill_value=0
    ).reset_index()

    # Ajouter les occurrences totales
    not_found_col = df_tag_pivot['exam_not_found'] if 'exam_not_found' in df_tag_pivot.columns else 0
    not_authorized_col = df_tag_pivot['exam_not_authorized'] if 'exam_not_authorized' in df_tag_pivot.columns else 0
    df_tag_pivot['Occurrences'] = not_found_col + not_authorized_col

    # Prendre le premier Id Externe pour chaque examen
    df_first_id = df_detailed_sorted.drop_duplicates(
        subset=['Examen Identifié', 'Catégorie']
    )[['Catégorie', 'Examen Identifié', 'Id Externe']]

    # Fusionner les données
    df_detailed_unique = df_tag_pivot.merge(df_first_id, on=['Catégorie', 'Examen Identifié'])

    # Réorganiser les colonnes pour avoir : Catégorie, Examen, Occurrences, Not Found, Not Authorized, Id Externe
    cols_order = ['Catégorie', 'Examen Identifié', 'Occurrences']
    if 'exam_not_found' in df_detailed_unique.columns:
        cols_order.append('exam_not_found')
    if 'exam_not_authorized' in df_detailed_unique.columns:
        cols_order.append('exam_not_authorized')
    cols_order.append('Id Externe')
    df_detailed_unique = df_detailed_unique[cols_order]

    # Trier par catégorie puis par occurrences décroissantes
    df_detailed_unique = df_detailed_unique.sort_values(
        ['Catégorie', 'Occurrences'],
        ascending=[True, False]
    )

    # Ajouter une colonne avec tous les Id Externes (pour les commentaires)
    df_detailed_unique['All_Id_Externes'] = df_detailed_unique.apply(
        lambda row: exam_ids_dict.get((row['Examen Identifié'], row['Catégorie']), []),
        axis=1
    )
    
    print("\n💾 Écriture du fichier Excel...")
    
    # Écrire dans Excel avec plusieurs onglets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Onglet Statistiques
        df_stats.to_excel(writer, sheet_name='Statistiques', index=False, startrow=1)
        
        # Onglets par catégorie (on garde temporairement All_Id_Externes pour le traitement)
        for category in sorted(df_detailed_unique['Catégorie'].unique()):
            df_cat = df_detailed_unique[df_detailed_unique['Catégorie'] == category].copy()
            sheet_name = category[:31]  # Excel limite à 31 caractères
            df_cat.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print("\n🎨 Application du design et ajout des graphiques...")
    
    # Rouvrir le fichier pour appliquer le design
    wb = load_workbook(output_file)
    
    # === DESIGN DE L'ONGLET STATISTIQUES ===
    ws_stats = wb['Statistiques']
    
    # Titre principal
    ws_stats['A1'] = 'Analyse des Appels Allodoct - Statistiques Globales'
    ws_stats['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_stats['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    ws_stats['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_stats.merge_cells('A1:E1')
    ws_stats.row_dimensions[1].height = 30
    
    # En-têtes du tableau (ligne 2)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col in range(1, 6):  # A à E (5 colonnes)
        cell = ws_stats.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Formater les données
    for row in range(3, ws_stats.max_row + 1):
        for col in range(1, 6):  # A à E (5 colonnes)
            cell = ws_stats.cell(row=row, column=col)

            # Bordures
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Alignement
            if col == 1 or col == 5:  # Catégorie et Top 5
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Coloration alternée des lignes
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

            # Mettre en gras la catégorie INTITULES INCOMPRIS en rouge
            if col == 1 and cell.value == 'INTITULES INCOMPRIS':
                cell.font = Font(bold=True, color='C00000')

    # Ajuster les largeurs de colonnes (A=Catégorie, B=Total, C=not_found, D=not_authorized, E=Tous les examens)
    ws_stats.column_dimensions['A'].width = 25  # Catégorie
    ws_stats.column_dimensions['B'].width = 12  # Total
    ws_stats.column_dimensions['C'].width = 18  # exam_not_found
    ws_stats.column_dimensions['D'].width = 20  # exam_not_authorized
    ws_stats.column_dimensions['E'].width = 60  # Tous les examens
    
    # === AJOUT DES GRAPHIQUES ACCESSIBLES ===
    
    # Position des graphiques
    chart_row = ws_stats.max_row + 3
    
    # 1. Camembert : Répartition par catégorie
    pie1 = PieChart()
    labels = Reference(ws_stats, min_col=1, min_row=3, max_row=ws_stats.max_row)
    data = Reference(ws_stats, min_col=2, min_row=2, max_row=ws_stats.max_row)
    pie1.add_data(data, titles_from_data=True)
    pie1.set_categories(labels)
    pie1.title = "Répartition des examens par catégorie"
    
    # Améliorer l'accessibilité : étiquettes de données visibles
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showCatName = True
    pie1.dataLabels.showVal = True
    pie1.dataLabels.showPercent = True
    pie1.dataLabels.separator = "\n"
    
    pie1.height = 12
    pie1.width = 18
    ws_stats.add_chart(pie1, f'A{chart_row}')

    # 2. Camembert : exam_not_found vs exam_not_authorized
    pie2 = PieChart()

    # Calculer les totaux globaux
    total_not_found = df_stats['exam_not_found'].sum()
    total_not_authorized = df_stats['exam_not_authorized'].sum()

    # Ajouter une mini table pour le graphique
    chart_data_row = ws_stats.max_row + 2
    ws_stats[f'A{chart_data_row}'] = 'Type'
    ws_stats[f'B{chart_data_row}'] = 'Nombre'
    ws_stats[f'A{chart_data_row + 1}'] = 'exam_not_found'
    ws_stats[f'B{chart_data_row + 1}'] = total_not_found
    ws_stats[f'A{chart_data_row + 2}'] = 'exam_not_authorized'
    ws_stats[f'B{chart_data_row + 2}'] = total_not_authorized

    # Masquer cette mini table
    for row in range(chart_data_row, chart_data_row + 3):
        ws_stats.row_dimensions[row].hidden = True

    data2 = Reference(ws_stats, min_col=2, min_row=chart_data_row, max_row=chart_data_row + 2)
    labels2 = Reference(ws_stats, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + 2)
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    pie2.title = "Répartition exam_not_found vs exam_not_authorized"

    # Améliorer l'accessibilité
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showCatName = True
    pie2.dataLabels.showVal = True
    pie2.dataLabels.showPercent = True
    pie2.dataLabels.separator = "\n"

    pie2.height = 12
    pie2.width = 18
    ws_stats.add_chart(pie2, f'K{chart_row}')
    
    # === DESIGN DES ONGLETS PAR CATÉGORIE ===
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Statistiques':
            continue
            
        ws = wb[sheet_name]
        
        # Identifier la colonne All_Id_Externes (dernière colonne)
        all_ids_col = ws.max_column
        id_externe_col = all_ids_col - 1  # Colonne Id Externe
        
        # En-têtes
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col in range(1, all_ids_col):  # Exclure la colonne All_Id_Externes
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Formater les données et ajouter les commentaires
        for row in range(2, ws.max_row + 1):
            # Récupérer la liste des Id Externes depuis la dernière colonne
            all_ids_cell = ws.cell(row=row, column=all_ids_col)
            all_ids_value = all_ids_cell.value
            
            # Parser la liste (stockée comme string)
            if all_ids_value and isinstance(all_ids_value, str):
                # Convertir la représentation string de liste en liste Python
                try:
                    import ast
                    all_ids_list = ast.literal_eval(all_ids_value)
                    if isinstance(all_ids_list, list) and len(all_ids_list) > 1:
                        # Créer le commentaire avec tous les Id Externes
                        comment_text = f"{len(all_ids_list)} Id Externes :\n\n" + "\n".join(all_ids_list)
                        
                        # Ajouter le commentaire sur la cellule "Occurrences" (colonne 3)
                        occurrences_cell = ws.cell(row=row, column=3)
                        comment = Comment(comment_text, "Allodoct Analysis")
                        comment.width = 400
                        comment.height = 200
                        occurrences_cell.comment = comment
                except:
                    pass
            
            # Formater toutes les colonnes sauf All_Id_Externes
            for col in range(1, all_ids_col):
                cell = ws.cell(row=row, column=col)
                
                # Bordures
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Alignement
                if col == 2:  # Examen Identifié
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Coloration alternée des lignes
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Supprimer la colonne All_Id_Externes
        ws.delete_cols(all_ids_col)

        # Ajuster les largeurs de colonnes (A=Catégorie, B=Examen Identifié, C=Occurrences, D=Not Found, E=Not Authorized, F=Id Externe)
        ws.column_dimensions['A'].width = 25  # Catégorie
        ws.column_dimensions['B'].width = 60  # Examen Identifié
        ws.column_dimensions['C'].width = 15  # Occurrences
        ws.column_dimensions['D'].width = 20  # exam_not_found
        ws.column_dimensions['E'].width = 20  # exam_not_authorized
        ws.column_dimensions['F'].width = 20  # Id Externe
        
        # Figer la première ligne
        ws.freeze_panes = 'A2'
    
    # Sauvegarder
    wb.save(output_file)
    
    print(f"\n✅ Analyse terminée ! Fichier créé: {output_file}")
    print(f"\n📊 Résumé:")
    print(f"   Total d'examens analysés: {len(df_detailed)}")
    print(f"   Examens uniques: {len(df_detailed_unique)}")
    print(f"   Catégories trouvées: {len(df_stats)}")
    print(f"   Intitulés incompris détectés: {len(df_detailed_unique[df_detailed_unique['Catégorie'] == 'INTITULES INCOMPRIS'])}")


if __name__ == '__main__':
    if len(sys.argv) != 5:
        print("Usage: python analyze_calls.py <not_found.xlsx> <not_authorized.xlsx> <reference.csv> <output.xlsx>")
        sys.exit(1)
    
    analyze_calls(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
